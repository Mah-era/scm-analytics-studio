"""
Offline Excel fixture generator snippets for SCM Analytics Studio.

XLSX: uses openpyxl if installed.
XLS: writes a small real binary Excel 2 BIFF .xls file using only the Python stdlib.
No external APIs are used.
"""

from pathlib import Path
import struct

def write_xlsx_with_openpyxl(path: str | Path, sheets: dict[str, list[list[object]]]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    header_fill = PatternFill("solid", fgColor="0F766E")
    header_font = Font(bold=True, color="FFFFFF")

    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name[:31])
        for row in rows:
            ws.append(row)
        if rows:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            ws.freeze_panes = "A2"
        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 35)

    wb.save(path)

def write_legacy_biff2_xls(path: str | Path, rows: list[list[object]]) -> None:
    def rec(opcode: int, data: bytes = b"") -> bytes:
        return struct.pack("<HH", opcode, len(data)) + data

    out = bytearray()
    out += rec(0x0009, struct.pack("<HH", 0x0002, 0x0010))  # BOF: Excel 2 worksheet
    out += rec(0x0000, struct.pack("<HHHH", 0, len(rows), 0, max(len(r) for r in rows)))  # DIMENSIONS
    xf = b"\x00\x00\x00"

    for r_idx, row in enumerate(rows):
        for c_idx, value in enumerate(row):
            if value is None:
                continue
            if isinstance(value, int) and -32768 <= value <= 32767:
                out += rec(0x0002, struct.pack("<HH", r_idx, c_idx) + xf + struct.pack("<h", value))
            elif isinstance(value, (float, int)):
                out += rec(0x0003, struct.pack("<HH", r_idx, c_idx) + xf + struct.pack("<d", float(value)))
            else:
                raw = str(value).encode("latin-1", errors="replace")[:255]
                out += rec(0x0004, struct.pack("<HH", r_idx, c_idx) + xf + struct.pack("B", len(raw)) + raw)

    out += rec(0x000A)  # EOF
    Path(path).write_bytes(out)

if __name__ == "__main__":
    rows = [
        ["date", "sku", "product", "demand_qty", "forecast_qty"],
        ["2026-01-01", "SKU-1001", "AeroBlend Mixer", 100, 105],
        ["2026-01-02", "SKU-1002", "Nimbus Rice Cooker", 95, 90],
    ]

    write_xlsx_with_openpyxl("example_multisheet.xlsx", {"Demand": rows, "HeaderOnly": [rows[0]], "Empty": []})
    write_legacy_biff2_xls("example_legacy.xls", rows)
